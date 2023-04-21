from queue import Queue
from tkinter import *
from tkinter import messagebox
import tkinter as tk
from tkinter import ttk
import pickle
import openpyxl
from openpyxl.utils import get_column_letter
import datetime
import pyttsx3

engine = pyttsx3.init("sapi5")
voices = engine.getProperty('voices')
engine.setProperty('voice', voices[2].id)
engine.setProperty('language', 'en-in')
engine.setProperty('gender', 'female')
engine.setProperty('volume', 1)
engine.setProperty('age', 'adult')
engine.setProperty('rate', 178)

def speak(Text):
    #print("    ")
    engine.say(text=Text)
    engine.runAndWait()
    #print("     ")


count = 0
item = ""
is_queue_running = False
first = True
row = 1
root = tk.Tk()
root.geometry("1650x850")
root.title("TOKEN GENERATOR")
root.resizable(False, False)
frame1 = Frame(root, highlightbackground="black", highlightthickness=3,width=1580, height=790, bd= 0)
frame1.place(x=33, y=26)

class QueueManager:
    global count
    def __init__(self):
        self.token_number = 0
        self.queue = []
        self.waiting = []
        self.tokens = " "
        self.total_count = 0
        

    def get_next_token(self):
        self.token_number += 1
        self.total_count += 1
        self.queue.append(self.token_number)
        global count
        count += 1

    def dequeue_token(self):
        if len(self.queue) > 0:
            global count
            count -= 1
            return self.queue.pop(0)
        else:
            return None
        

queue_manager = QueueManager()

# Load saved state if available
try:
    with open("state.pickle", "rb") as f:
        saved_state = pickle.load(f)
        count = saved_state["count"]
        row = saved_state["row"]
        queue_manager.tokens = saved_state["tokens"]
        queue_manager.token_number = saved_state["token_number"]
        queue_manager.queue = saved_state["queue"]
        queue_manager.waiting = saved_state["waiting"]
        
except FileNotFoundError:
    # No saved state found
    pass

def save_state():
    with open("state.pickle", "wb") as f:
        saved_state = {
            "count": count,
            "row":row,
            "tokens": queue_manager.tokens,
            "token_number": queue_manager.token_number,
            "queue": queue_manager.queue,
            "waiting": queue_manager.waiting,
            
        }
        pickle.dump(saved_state, f)

def close_window():
    save_state()
    root.destroy()

# Save state when window is closed
root.protocol("WM_DELETE_WINDOW", close_window)

#workbook = openpyxl.Workbook()
workbook = openpyxl.load_workbook(filename='Tokens Generated.xlsx')
worksheet = workbook.active
'''
worksheet['A1'] = 'DATE'
worksheet['B1'] = 'TIME'
worksheet['C1'] = 'TOKENS GENERATED'
'''
row = worksheet.max_row + 1


token_count = tk.Label(root,text = " ", font=("Arial", 17),fg="black", bg="white",width=40,height=1)
token_count.place(x=100,y=200)
token_count.config(text=f"REMAINING NUMBER OF TOKENS: {count}")

def generate_token():
    queue_manager.get_next_token()
    token_count.config(text=f"REMAINING NUMBER OF TOKENS: {count}")
    #print(count)
    tokens = " "
    for i in queue_manager.queue:
        if tokens == " ":
            tokens = tokens + str(i)
        else:
            tokens = tokens + ", " + str(i)
    generated_tokens.delete("0",tk.END)
    generated_tokens.insert(tk.END,f"{tokens}")
    queue_manager.tokens = tokens


def start_screen():
    global is_queue_running
    is_queue_running = True
    global first
    tokens = " "
    if len(queue_manager.queue) > 0 and first:
        global item
        first = False
        item = queue_manager.dequeue_token()
        #print(item)
        screen.config(text=f"{item}")
        speak(f"Token Number {item}")
        token_count.config(text=f"REMAINING NUMBER OF TOKENS: {count}")
        for i in queue_manager.queue:
            if tokens == " ":
                tokens = tokens +str(i)
            else:
                tokens = tokens + ", " + str(i)
        generated_tokens.delete("0",tk.END)
        generated_tokens.insert(tk.END,f"{tokens}")
        queue_manager.tokens = tokens

def next_token_screen():
    if is_queue_running:
        global item
        tokens = " "
        item = queue_manager.dequeue_token()
        
        #print(item)
        if item == None:
            to_queue()
        else:
            screen.config(text=f"{item}")
            speak(f"Token Number {item}")
        token_count.config(text=f"REMAINING NUMBER OF TOKENS: {count}")
        for i in queue_manager.queue:
            if tokens == " ":
                tokens = tokens +str(i)
            else:
                tokens = tokens + ", " + str(i)
        generated_tokens.delete("0",tk.END)
        generated_tokens.insert(tk.END,f"{tokens}")
        queue_manager.tokens = tokens

def to_queue():
    global count
    global item
    count_queue = 0
    queue_manager.queue.clear()
    for i in range(waiting_list.size()):
        
        count_queue = count_queue + 1
        queue_manager.queue.append(waiting_list.get(i))
        #print(to_queue)
        #print(count_queue)
        #print(queue_manager.queue)
    queue_manager.waiting = []
    waiting_list.delete(0,tk.END)
    count = count_queue
    item = queue_manager.dequeue_token()
    screen.config(text=f"{item}")
    speak(f"Token Number {item}")
    
        
def stop_screen():
    global is_queue_running
    global first
    first = True
    is_queue_running = False

def add_to_wait():
    global is_queue_running
    global item
    tokens = " "
    if is_queue_running:
        queue_manager.waiting.append(item)
        #print(queue_manager.waiting)
        waiting_list.insert(END,item)
        item = queue_manager.dequeue_token()
        if item == None:
            to_queue()
        else:
            screen.config(text=f"{item}")
            speak(f"Token Number {item}")
        token_count.config(text=f"REMAINING NUMBER OF TOKENS: {count}")
        for i in queue_manager.queue:
            if tokens == " ":
                tokens = tokens +str(i)
            else:
                tokens = tokens + ", " + str(i)
        generated_tokens.delete("0",tk.END)
        generated_tokens.insert(tk.END,f"{tokens}")
        queue_manager.tokens = tokens

def screen_insert():
    global item
    #print(item)
    item = waiting_list.get(ANCHOR)
    #print(item1)
    tokens = " "
    queue_manager.waiting.remove(item)
    #print(queue_manager.waiting)
    screen.config(text=f"{item}")
    speak(f"Token Number {item}")
    token_count.config(text=f"REMAINING NUMBER OF TOKENS: {count}")
    for i in queue_manager.queue:
        if tokens == " ":
            tokens = tokens +str(i)
        else:
            tokens = tokens + ", " + str(i)
    generated_tokens.delete("0",tk.END)
    generated_tokens.insert(tk.END,f"{tokens}")
    queue_manager.tokens = tokens
    waiting_list.delete(ANCHOR)
    
#reset tokens
def token_reset():
    global count
    global row
    column = 0
    total_count = queue_manager.total_count
    count = 0
    queue_manager.__init__()
    token_count.config(text=f"REMAINING NUMBER OF TOKENS: {count}")
    generated_tokens.delete("0",tk.END)
    screen.config(text=" ")
    waiting_list.delete(0,tk.END)
    now = datetime.datetime.now()
    current_date = now.strftime("%d/%m/%Y")
    #print(current_date)
    current_time = now.strftime("%H:%M:%S")
    #print(current_time)
    
    if total_count != 0:
        worksheet.cell(row=row, column=1, value=current_date)
        worksheet.cell(row=row, column=2, value=current_time)
        worksheet.cell(row=row, column=3, value=total_count)
        # Increment the row index
        row += 1
        # Save the workbook to a file
        workbook.save('Tokens Generated.xlsx')
#print(row)

    


style = ttk.Style()
style.theme_use("default")
style.configure("Vertical.TScrollbar", gripcount=0, background="#f0f0f0", troughcolor="#d9d9d9", bordercolor="#d9d9d9", darkcolor="#d9d9d9", lightcolor="#d9d9d9")
scrollbar = ttk.Scrollbar(root, style="Vertical.TScrollbar")
style.configure("Horizontal.TScrollbar", gripcount=0, background="#f0f0f0", troughcolor="#d9d9d9", bordercolor="#d9d9d9", darkcolor="#d9d9d9", lightcolor="#d9d9d9")
hscrollbar = ttk.Scrollbar(root, style="Horizontal.TScrollbar", orient=tk.HORIZONTAL)

screen = tk.Label(root, text="", font=("Arial", 30), bg="black", fg="white", width=19, height=5)
screen.place(x=750,y=270)

start_button = tk.Button(root, text="START", font=("Arial", 16), command=start_screen, width=7)
start_button.place(x=100,y=320)

next_token = tk.Button(root, text="NEXT TOKEN", font=("Arial", 16), command=next_token_screen, width=14)
next_token.place(x=240,y=350)

stop_button = tk.Button(root, text="STOP", font=("Arial", 16), command=stop_screen, width=7)
stop_button.place(x=100,y=380)

waiting_status = tk.Label(root,text = " ", font=("Arial", 18),fg="black",width=25,height=2)
waiting_status.place(x=780,y=540)
waiting_status.config(text="WAITING LIST")

waiting_list = tk.Listbox(root, font="Arial 16",  width=37, height=8)
waiting_list.config(yscrollcommand=scrollbar.set)
scrollbar.config(command=waiting_list.yview)
waiting_list.place(x=750,y=600)
scrollbar.place(x=1198, y=600, height=204)
for i in queue_manager.waiting:
    waiting_list.insert(tk.END,i)

#print(queue_manager.waiting)

generated = tk.Label(root,text = " ", font=("Arial", 18),fg="black",width=25,height=1)
generated.config(text="GENERATED TOKENS:")
generated.place(x=750,y=100)

generated_tokens = tk.Listbox(root,font=("Arial", 18),fg="black",bg="white",width=34,height=1)
generated_tokens.place(x=750,y=170)
generated_tokens.config(xscrollcommand=hscrollbar.set)
hscrollbar.config(command=generated_tokens.xview)
hscrollbar.place(x=750, y=200, width=445, height=20)
generated_tokens.delete("0",tk.END)
generated_tokens.insert(tk.END,f"{queue_manager.tokens}")

wait_button = tk.Button(root, text="WAIT", font=("Arial", 16), command=add_to_wait, width=7)
wait_button.place(x=100,y=640)

insert_button = tk.Button(root, text="INSERT TO SCREEN", font=("Arial", 16), command=screen_insert)
insert_button.place(x=100,y=700)

generate_button = tk.Button(root, text="GENERATE TOKENS", font=("Arial", 16), command=generate_token, height=2)
generate_button.place(x=1360,y=700)

reset_button = tk.Button(root,text="RESET", font=("Arial",16),width=10, height=2, command=token_reset)
reset_button.place(x=1400,y=600)


root.mainloop()